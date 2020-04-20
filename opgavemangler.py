# lavet af Nikolaj Jehøj-Krogager

output = {"results": [], "errors": [], "debug": [], "name": ""}

try:    
    import os
    from pathlib import Path
    import json
    from openpyxl import load_workbook
except Exception as e:
    # FEJL - openpyxl er ikke installeret
    output["errors"].append("Mangler openpyxl")
    output["debug"].append(e)
    print(json.dumps(output))
    quit()

def checkFileInCorrectPlace():
    # Check if folder 4/5 parents up equals AARHUS TECH
    foldername = os.path.basename(Path(__file__).parent.parent.parent.parent.parent)
    return foldername == "AARHUS TECH"

def returnOpgaver(pathToOpgaveOversigt):
    wb = load_workbook(filename = pathToOpgaveOversigt)
    sheet_ranges = wb["Ark1"]
    kapDict = {}
    
    for row in range(sheet_ranges.max_row):
        kap = sheet_ranges["A" + str(row + 1)].value
        if kap is None:
            break
        
        value = sheet_ranges["B" + str(row + 1)].value
        value = value.split(",")
        # print(value)

        opgaver = []
        
        for opgave in value:
            opgave = opgave.split("-")
            if type(opgave) == list and len(opgave) == 2:
                for element in list(range(int(opgave[0]), int(opgave[1]) + 1)):
                    opgaver.append(str(element).zfill(3) + "a")
            else:
                opgaver.append(str(*opgave).zfill(3) + "a")

        kapDict[kap] = opgaver
    return kapDict

def getOwnFiles(path, kapitler):
    ownFiles = {}
    temp = os.listdir(path)
    for item in range(len(temp)):
        if temp[item] != "Mirsad Kadribasic - ZZMatInfoMappe" and temp[item][:6] == "Mirsad":
            path = os.path.join(path, temp[item])
            break
    
    for x in range(len(kapitler)):
        ownFiles[kapitler[x]] = []
        tempPath = "{}/{}".format(path, kapitler[x])
        try:
            opgaver = os.listdir(tempPath) 
        except FileNotFoundError as error:
            # der mangler en mappe
            output["errors"].append(f"Du mangler en mappe {error}")
            # fjern mappen fra listen (eller stop programmet)
            print(json.dumps(output))
            quit()
        for y in range(len(opgaver)):
            if opgaver[y][-4:] == ".pdf" and opgaver[y][0] != ".":
                ownFiles[kapitler[x]].append(opgaver[y][:4])
    return ownFiles

def compare(opgaver, egneOpgaver):
    manglende = {}
    m = []
    for x in opgaver.items():
        kap = x[0]
        opg = x[1]
        manglende[kap] = []
        for i in range(len(opg)):
            if opg[i] in egneOpgaver[kap]:
                continue
                #print(opg[i])
            else:
                manglende[kap].append(opg[i])
                m.append("{}: {}".format(kap, opg[i]))
    if len(m) <= 0:
        x = True
    else:
        x = False
    return manglende, x, m

if __name__ == "__main__":
    output["debug"].append(f"Filen er placeret her: {__file__}")

    if checkFileInCorrectPlace():
        output["name"] = f"{os.path.basename(Path(__file__).parent.parent.parent.parent)[20:]}"
    else:
        output["errors"].append("Python-filen er placeret det forkerte sted. Den skal placeres i Aarhus Tech under dit navn!")
        print(json.dumps(output, ensure_ascii=False))
        quit()

    aarhusTechFolderPath = Path(__file__).parent.parent.parent.parent.parent
    pathToOpgaver = os.path.join(aarhusTechFolderPath, 'Mirsad Kadribasic - ZZMatInfoMappe', 'Opgaver18R.xlsx')
        
    mirsadOpgaver = returnOpgaver(pathToOpgaver)
    kapitler = list(mirsadOpgaver.keys())
    dineOpgaverPath = Path(__file__).parent.parent.parent.parent
    dineOpgaver = getOwnFiles(str(__file__)[2].join(dineOpgaverPath), kapitler)

    comparison, x, m = compare(mirsadOpgaver, dineOpgaver)

    for kap in comparison:
        if len(comparison[kap]) == 0:
            pass
        else:
            result = [kap, []]
            temp = ', '.join(comparison[kap])
            result[1].append(temp)
            output["results"].append(result)
    
    done = True
    for kap in comparison:
        if len(comparison[kap]) != 0:
            done = False
            break
    
    if done:
        output["results"].append(["Super!", "Du mangler ingen opgaver!"])
    
    print(json.dumps(output))
